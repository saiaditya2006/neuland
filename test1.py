# import pandas as pd
# from fuzzywuzzy import fuzz
# import re
# from collections import Counter
# from openpyxl import Workbook
# from openpyxl.utils.dataframe import dataframe_to_rows
# from openpyxl.styles import Font

# # Load the Excel file
# df = pd.read_excel("neudata.xlsx")

# # Define stopwords to clean
# company_stopwords = [
#     r'\bco\b', r'\bltd\b', r'\binc\b', r'\bproducts\b', r'\bchas\b', r'\band\b'
#     r'\bcorporation\b', r'\bcorp\b', r'\bprivate\b', r'\blimited\b',
#     r'\bplc\b', r'\bpvt\b', r'\bgroup\b', r'\bcompany\b', r'\bchemical\b', r'\bchemicals', r'laboratories'
# ]

# # Function to clean company names
# def clean_company_name(name):
#     name = name.lower()
#     name = re.sub('|'.join(company_stopwords), '', name)
#     name = re.sub(r'[^a-z0-9\s]', '', name)
#     name = re.sub(r'\s+', ' ', name).strip()
#     return name

# # Clean and prepare data
# df['HOLDER'] = df['HOLDER'].astype(str)
# df['HOLDER_clean'] = df['HOLDER'].apply(clean_company_name)

# # Create clusters
# clusters = []
# cluster_map = {}
# threshold = 80

# for val_i in df['HOLDER_clean']:
#     matched = False
#     for idx, cluster in enumerate(clusters):
#         for val_j in cluster:
#             if fuzz.token_sort_ratio(val_i, val_j) >= threshold:
#                 cluster.append(val_i)
#                 cluster_map[val_i] = idx
#                 matched = True
#                 break
#         if matched:
#             break
#     if not matched:
#         clusters.append([val_i])
#         cluster_map[val_i] = len(clusters) - 1

# # Map most frequent original name in each cluster
# cluster_representatives = {}

# for idx, cluster in enumerate(clusters):
#     original_names = df[df['HOLDER_clean'].isin(cluster)]['HOLDER']
#     most_common_original = Counter(original_names).most_common(1)[0][0]
#     cluster_representatives[idx] = most_common_original

# # Apply standardized name
# def get_standardized_name(clean_name):
#     cluster_idx = cluster_map.get(clean_name)
#     return cluster_representatives.get(cluster_idx, clean_name)

# df['HOLDER_final'] = df['HOLDER_clean'].apply(get_standardized_name)

# # Create Excel workbook and sheet
# wb = Workbook()
# ws = wb.active
# ws.title = "Standardized Companies"

# # Add DataFrame rows to sheet
# for r_idx, row in enumerate(dataframe_to_rows(df[['HOLDER', 'HOLDER_final']], index=False, header=True), 1):
#     for c_idx, value in enumerate(row, 1):
#         cell = ws.cell(row=r_idx, column=c_idx, value=value)
#         # Bold the header
#         if r_idx == 1:
#             cell.font = Font(bold=True)

# # Adjust column widths
# for col in ws.columns:
#     max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
#     col_letter = col[0].column_letter
#     ws.column_dimensions[col_letter].width = max_length + 5  # add padding

# # Save the workbook
# wb.save("standardized_companies.xlsx")
# print("✅ Excel file 'standardized_companies.xlsx' created with formatting.")


import pandas as pd
from fuzzywuzzy import fuzz
import re
from collections import Counter
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font

# Load the Excel file
df = pd.read_excel("neudata.xlsx")

# Define stopwords to clean
company_stopwords = [
    r'\bco\b', r'\bltd\b', r'\binc\b', r'\bproducts\b', r'\bchas\b', r'\band\b',
    r'\bcorporation\b', r'\bcorp\b', r'\bprivate\b', r'\blimited\b',
    r'\bplc\b', r'\bpvt\b', r'\bgroup\b', r'\bcompany\b', r'\bchemical\b', r'\bchemicals\b', r'\blaboratories\b'
]

# Function to clean company names
def clean_company_name(name):
    name = name.lower()
    name = re.sub('|'.join(company_stopwords), '', name)
    name = re.sub(r'[^a-z0-9\s]', '', name)
    name = re.sub(r'\s+', ' ', name).strip()
    return name

# Clean and prepare data
df['HOLDER'] = df['HOLDER'].astype(str)
df['HOLDER_clean'] = df['HOLDER'].apply(clean_company_name)

# Create clusters
clusters = []
cluster_map = {}
threshold = 80

for val_i in df['HOLDER_clean']:
    matched = False
    for idx, cluster in enumerate(clusters):
        for val_j in cluster:
            if fuzz.token_sort_ratio(val_i, val_j) >= threshold:
                cluster.append(val_i)
                cluster_map[val_i] = idx
                matched = True
                break
        if matched:
            break
    if not matched:
        clusters.append([val_i])
        cluster_map[val_i] = len(clusters) - 1

# Map most frequent original name in each cluster
cluster_representatives = {}
for idx, cluster in enumerate(clusters):
    original_names = df[df['HOLDER_clean'].isin(cluster)]['HOLDER']
    most_common_original = Counter(original_names).most_common(1)[0][0]
    cluster_representatives[idx] = most_common_original

# Apply standardized name (parent company)
def get_standardized_name(clean_name):
    cluster_idx = cluster_map.get(clean_name)
    return cluster_representatives.get(cluster_idx, clean_name)

df['PARENT_COMPANY'] = df['HOLDER_clean'].apply(get_standardized_name)

# Create Excel workbook and sheet
wb = Workbook()
ws = wb.active
ws.title = "Standardized Companies"

# Add DataFrame rows to sheet (HOLDER, HOLDER_clean, PARENT_COMPANY)
output_cols = ['HOLDER', 'HOLDER_clean', 'PARENT_COMPANY']
for r_idx, row in enumerate(dataframe_to_rows(df[output_cols], index=False, header=True), 1):
    for c_idx, value in enumerate(row, 1):
        cell = ws.cell(row=r_idx, column=c_idx, value=value)
        # Bold the header
        if r_idx == 1:
            cell.font = Font(bold=True)

# Adjust column widths
for col in ws.columns:
    max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
    col_letter = col[0].column_letter
    ws.column_dimensions[col_letter].width = max_length + 5  # add padding

# Save the workbook
wb.save("standardized_companies.xlsx")
print("✅ Excel file 'standardized_companies.xlsx' created with parent company mapping.")


