import pandas as pd
from fuzzywuzzy import process, fuzz
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

# Load the Excel file
df = pd.read_excel("neudata.xlsx")

# Step 1: Clean and standardize
df['HOLDER'] = df['HOLDER'].astype(str)
df['HOLDER_clean'] = df['HOLDER'].str.lower().str.strip()

# Create HOLDER_CLEANED as the working column (initially same as HOLDER)
df['HOLDER_CLEANED'] = df['HOLDER']

# Step 2: Define fuzzy matching and replacement function
def find_and_replace_multiple_companies(df, replacements, threshold=80):
    for search_term, new_name in replacements.items():
        search_term = search_term.lower().strip()
        unique_companies = df['HOLDER_clean'].unique()

        # Fuzzy and substring match
        matches = process.extract(search_term, unique_companies, scorer=fuzz.token_sort_ratio)
        matched_clean_names = [match[0] for match in matches if match[1] >= threshold or search_term in match[0]]

        # Replace only in HOLDER_CLEANED
        df['HOLDER_CLEANED'] = df.apply(
            lambda row: new_name if row['HOLDER_clean'] in matched_clean_names or search_term in row['HOLDER_clean']
            else row['HOLDER_CLEANED'],
            axis=1
        )
    
    return df

# Step 3: Replacements dictionary (can be user input too)
replacements = {
    'pfizer': 'Pfizer Inc',
    'johnson': 'Johnson & Johnson',
    'novartis': 'Novartis AG',
}

# Apply replacements
df_cleaned = find_and_replace_multiple_companies(df, replacements)

# Drop the internal 'HOLDER_clean' column
df_cleaned = df_cleaned.drop(columns=['HOLDER_clean'])

# Reorder columns to have HOLDER and HOLDER_CLEANED side by side
cols = df_cleaned.columns.tolist()

# Ensure correct ordering: move HOLDER and HOLDER_CLEANED to the front, in order
cols.remove('HOLDER')
cols.remove('HOLDER_CLEANED')
ordered_cols = ['HOLDER', 'HOLDER_CLEANED'] + cols

# Create a new Excel workbook and worksheet
wb = Workbook()
ws = wb.active
ws.title = "Cleaned Data"

# Write the reordered DataFrame to the worksheet
for r in dataframe_to_rows(df_cleaned[ordered_cols], index=False, header=True):
    ws.append(r)

# Auto-adjust column widths
for col in ws.columns:
    max_length = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        if cell.value:
            max_length = max(max_length, len(str(cell.value)))
    ws.column_dimensions[col_letter].width = max_length + 2

# Save the file
output_path = "neudata_final_cleaned.xlsx"
wb.save(output_path)

print(f"✅ Excel file saved with HOLDER and HOLDER_CLEANED side by side: {output_path}")
