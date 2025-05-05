# Re-import necessary modules after code execution environment reset
import pandas as pd
from fuzzywuzzy import fuzz
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

# Reload the uploaded Excel file
file_path = "neudata.xlsx"
df = pd.read_excel(file_path)

# Step 1: Clean HOLDER column
df['HOLDER'] = df['HOLDER'].astype(str)
df['HOLDER_clean'] = df['HOLDER'].str.lower().str.strip()

# Step 2: Group similar names using fuzzy matching
unique_names = df['HOLDER_clean'].unique()
group_map = {}
group_counts = defaultdict(int)

for i, name in enumerate(unique_names):
    if name in group_map:
        continue
    group_map[name] = name  # Initial assignment
    for other in unique_names[i + 1:]:
        if other in group_map:
            continue
        if fuzz.token_sort_ratio(name, other) >= 85:
            group_map[other] = name

# Count occurrences of each cleaned name
for original_name in df['HOLDER_clean']:
    representative = group_map[original_name]
    group_counts[representative] += 1

# Determine the parent (most frequent) name for each group
reverse_group_map = defaultdict(list)
for k, v in group_map.items():
    reverse_group_map[v].append(k)

final_parent_map = {}
for rep_name, members in reverse_group_map.items():
    sub_df = df[df['HOLDER_clean'].isin(members)]
    most_common_name = sub_df['HOLDER_clean'].value_counts().idxmax()
    for name in members:
        final_parent_map[name] = most_common_name

# Step 3: Add Parent_Company column
df['Parent_Company'] = df['HOLDER_clean'].map(final_parent_map)

# Step 4: Save to Excel using openpyxl with formatting
wb = Workbook()
ws = wb.active
ws.title = "Cleaned Data"
df_to_save = df.drop(columns=['HOLDER_clean'])

for r in dataframe_to_rows(df_to_save, index=False, header=True):
    ws.append(r)

# Auto-adjust column widths
for col in ws.columns:
    max_length = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        try:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        except:
            pass
    ws.column_dimensions[col_letter].width = max_length + 2

output_excel = "neudata_with_parent_companies.xlsx"
wb.save(output_excel)

output_excel
