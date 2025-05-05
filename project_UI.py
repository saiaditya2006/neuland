import streamlit as st
import pandas as pd
import json
from fuzzywuzzy import process, fuzz
from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

st.title("🧹 Company Name Cleaner with Fuzzy Matching")

# Upload Excel file
uploaded_file = st.file_uploader("📂 Upload your Excel file (.xlsx)", type=["xlsx"])
threshold = 80

# Replacement dictionary input
user_input = st.text_area("✍️ Enter replacements as JSON (e.g., {\"pfizer\": \"Pfizer Inc\"})")

# Parse user input
replacements = {}
if user_input:
    try:
        replacements = json.loads(user_input)
    except json.JSONDecodeError:
        st.error("❌ Invalid JSON input. Please enter a valid JSON dictionary like: {\"pfizer\": \"Pfizer Inc\"}")

# Show data preview if file is uploaded
df = None
if uploaded_file:
    try:
        df = pd.read_excel(uploaded_file)
        st.write("📊 First few rows of the uploaded file:")
        st.dataframe(df.head())
    except Exception as e:
        st.error(f"❌ Error reading Excel file: {e}")
        df = None

# Show run button always if file and replacements input exist
if st.button("🚀 Run Cleaning"):
    if not uploaded_file:
        st.warning("⚠️ Please upload an Excel file first.")
    elif not replacements:
        st.warning("⚠️ Please provide a valid replacement dictionary in JSON format.")
    elif df is not None:
        if 'HOLDER' not in df.columns:
            st.error("❌ The uploaded file does not contain a column named 'HOLDER'.")
        else:
            # Begin processing
            df['HOLDER'] = df['HOLDER'].astype(str)
            df['HOLDER_clean'] = df['HOLDER'].str.lower().str.strip()
            df['HOLDER_CLEANED'] = df['HOLDER']  # Copy original

            for search_term, new_name in replacements.items():
                search_term = search_term.lower().strip()
                unique_companies = df['HOLDER_clean'].unique()

                matches = process.extract(search_term, unique_companies, scorer=fuzz.token_sort_ratio)
                matched_clean_names = [match[0] for match in matches if match[1] >= threshold or search_term in match[0]]

                df['HOLDER_CLEANED'] = df.apply(
                    lambda row: new_name if row['HOLDER_clean'] in matched_clean_names or search_term in row['HOLDER_clean']
                    else row['HOLDER_CLEANED'],
                    axis=1
                )

            df = df.drop(columns=['HOLDER_clean'])

            # Reorder columns
            cols = df.columns.tolist()
            cols.remove('HOLDER')
            cols.remove('HOLDER_CLEANED')
            df = df[['HOLDER', 'HOLDER_CLEANED'] + cols]

            # Save to Excel in memory
            output = BytesIO()
            wb = Workbook()
            ws = wb.active
            ws.title = "Cleaned Data"
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = max_len + 2
            wb.save(output)

            st.success("✅ Cleaning complete.")
            st.download_button("📥 Download Cleaned Excel", output.getvalue(), file_name="cleaned_data.xlsx")
