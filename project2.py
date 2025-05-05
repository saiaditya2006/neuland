# Re-run the user's preferred fuzzy replacement method and export using openpyxl with formatting
import pandas as pd
from fuzzywuzzy import process, fuzz
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

# Load the Excel file
df = pd.read_excel("neudata.xlsx")

# Step 1: Clean and standardize for fuzzy matching
df['HOLDER'] = df['HOLDER'].astype(str)
df['HOLDER_clean'] = df['HOLDER'].str.lower().str.strip()
df['HOLDER_standardized'] = df['HOLDER_clean']

# Step 2: Define fuzzy matching and replacement
def find_and_replace_partial_company(df, search_term, new_name=None, threshold=80):
    search_term = search_term.lower().strip()
    unique_companies = df['HOLDER_clean'].unique()

    # Fuzzy and substring match
    matches = process.extract(search_term, unique_companies, scorer=fuzz.token_sort_ratio)
    matched_clean_names = [match[0] for match in matches if match[1] >= threshold or search_term in match[0]]

    # Replace with new name if specified
    if new_name:
        df['HOLDER'] = df.apply(
            lambda row: new_name if row['HOLDER_clean'] in matched_clean_names or search_term in row['HOLDER_clean']
            else row['HOLDER'],
            axis=1
        )
        df['HOLDER_clean'] = df['HOLDER'].str.lower().str.strip()
        df['HOLDER_standardized'] = df['HOLDER_clean']
    
    return df

# Step 3: Apply replacement for 'pfizer'
df_cleaned = find_and_replace_partial_company(df, search_term='pfizer', new_name='Pfizer Inc')

# Step 4: Write to Excel with auto column width formatting
wb = Workbook()
ws = wb.active
ws.title = "Cleaned Data"

for r in dataframe_to_rows(df_cleaned.drop(columns=['HOLDER_clean', 'HOLDER_standardized']), index=False, header=True):
    ws.append(r)

for col in ws.columns:
    max_length = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        try:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        except:
            pass
    ws.column_dimensions[col_letter].width = max_length + 2

# Save the cleaned and formatted Excel file
output_path = "neudata_standardized_autowidth2.xlsx"
wb.save(output_path)

output_path


# import pandas as pd
# from fuzzywuzzy import process, fuzz
# from openpyxl import Workbook
# from openpyxl.utils.dataframe import dataframe_to_rows
# from openpyxl.utils import get_column_letter

# # Load the Excel file
# df = pd.read_excel("neudata.xlsx")

# # Generic fuzzy match + replace function for any column
# def find_and_replace_partial(df, column, search_term, new_name=None, threshold=80):
#     search_term = search_term.lower().strip()
#     df[f'{column}_clean'] = df[column].astype(str).str.lower().str.strip()
#     df[f'{column}_standardized'] = df[f'{column}_clean']
    
#     unique_values = df[f'{column}_clean'].unique()
#     matches = process.extract(search_term, unique_values, scorer=fuzz.token_sort_ratio)
#     matched_clean_names = [match[0] for match in matches if match[1] >= threshold or search_term in match[0]]

#     if new_name:
#         df[column] = df.apply(
#             lambda row: new_name if row[f'{column}_clean'] in matched_clean_names or search_term in row[f'{column}_clean']
#             else row[column],
#             axis=1
#         )
#         df[f'{column}_clean'] = df[column].astype(str).str.lower().str.strip()
#         df[f'{column}_standardized'] = df[f'{column}_clean']
    
#     return df

# # === Apply to HOLDER column ===
# df = find_and_replace_partial(df, column='HOLDER', search_term='pfizer', new_name='Pfizer Inc')

# # === Apply to SUBJECT column ===
# df = find_and_replace_partial(df, column='SUBJECT', search_term='penicillin', new_name='Penicillin')

# # === Write to Excel using openpyxl with auto column width ===
# wb = Workbook()
# ws = wb.active
# ws.title = "Cleaned Data"

# # Drop helper columns before saving
# df_to_save = df.drop(columns=[col for col in df.columns if col.endswith('_clean') or col.endswith('_standardized')])

# for r in dataframe_to_rows(df_to_save, index=False, header=True):
#     ws.append(r)

# # Auto-adjust column widths
# for col in ws.columns:
#     max_length = 0
#     col_letter = get_column_letter(col[0].column)
#     for cell in col:
#         try:
#             if cell.value:
#                 max_length = max(max_length, len(str(cell.value)))
#         except:
#             pass
#     ws.column_dimensions[col_letter].width = max_length + 2

# # Save the output
# output_path = "neudata_standardized_final.xlsx"
# wb.save(output_path)

# print(f"File saved to: {output_path}")
