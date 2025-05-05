import pandas as pd
from fuzzywuzzy import fuzz
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

# Load the data
df = pd.read_excel("neudata.xlsx")

# Clean the HOLDER column
df['HOLDER'] = df['HOLDER'].astype(str)
df['HOLDER_clean'] = df['HOLDER'].str.lower().str.strip()

# Step 1: Group similar company names using fuzzy matching
unique_names = df['HOLDER_clean'].unique()
groups = []
used = set()
threshold = 80

for name in unique_names:
    if name in used:
        continue
    group = [name]
    used.add(name)
    for other in unique_names:
        if other in used:
            continue
        if fuzz.token_sort_ratio(name, other) >= threshold:
            group.append(other)
            used.add(other)
    groups.append(group)

# Step 2: For each group, find the most frequent representative name
name_to_parent = {}
for group in groups:
    # Count total occurrences of each name in this group
    counts = df['HOLDER_clean'].value_counts()
    representative = max(group, key=lambda x: counts.get(x, 0))
    for name in group:
        name_to_parent[name] = representative.title()  # Capitalize nicely

# Step 3: Create the new PARENT_COMPANY column
df['PARENT_COMPANY'] = df['HOLDER_clean'].map(name_to_parent)

# Step 4: Drop helper columns and save to Excel with formatting
df_final = df.drop(columns=['HOLDER_clean'])

wb = Workbook()
ws = wb.active
ws.title = "Standardized Data"

for r in dataframe_to_rows(df_final, index=False, header=True):
    ws.append(r)

# Auto-adjust column widths
for col in ws.columns:
    max_len = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        try:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        except:
            pass
    ws.column_dimensions[col_letter].width = max_len + 2

# Save file
output_path = "neudata_with_parent_company.xlsx"
wb.save(output_path)

print(f"✅ File saved to: {output_path}")
